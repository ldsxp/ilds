# -*- coding: utf-8 -*-
#
# ---------------------------------------
#   程序：excel_xlsx.py
#   版本：0.6
#   作者：lds
#   日期：2024-08-29
#   语言：Python 3.X
#   说明：读取excel文件内容
# ---------------------------------------

import os
from time import time
from warnings import warn
from pathlib import Path

from .file import get_dir_files

# ReadExcel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, is_date_format

(
    XL_CELL_EMPTY,
    XL_CELL_TEXT,
    XL_CELL_NUMBER,
    XL_CELL_DATE,
    XL_CELL_BOOLEAN,
    XL_CELL_ERROR,
    XL_CELL_BLANK,
) = range(7)


# warn('我们把 excel_xlsx 重命名为 excel，以后会删除 excel_xlsx', DeprecationWarning, stacklevel=2)


def get_title_style(workbook, style_colour='#daeef3', is_bold=False):
    """
    获取自定义颜色的标题风格
    :param workbook:
    :param style_colour: 默认颜色为什么浅色
    :param is_bold: 是否设置粗体字体
    :return:
    """
    head_style = workbook.add_format()
    # 设置粗体
    if is_bold:
        head_style.set_bold()
    # 设置对齐
    head_style.set_align('center')
    head_style.set_align('vcenter')
    # 设置颜色
    head_style.set_bg_color(style_colour)
    # 设置边框
    head_style.set_border()
    # 调整行的高度以适应文本
    head_style.set_text_wrap()
    return head_style


def get_excel_info(file, only_read_first_table=False):
    """
    读取 Excel 信息

    :param file: Excel 文件
    :param only_read_first_table: 只读取第一个表格
    :return: {sheet_name: {'file_name', 'index', 'sheet_name', 'sheet_names', 'max_row', 'max_column', 'columns'}, }
    """
    data = {}

    file_name = os.path.basename(file)

    wb = load_workbook(file, read_only=True)
    sheet_names = wb.sheetnames

    for index, sheet_name in enumerate(sheet_names):
        ws = wb[sheet_name]

        # 获取第一行数据
        columns = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        df_data = {
            'file_name': file_name,
            'index': index,
            'sheet_name': sheet_name,
            'sheet_names': sheet_names,
            'sheet_state': ws.sheet_state,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'columns': list(columns),

        }

        # 我们添加 count 为了和以前获取信息相同，以后会去掉，用 max_row 替代
        # df_data['count'] = df_data['max_row']

        if only_read_first_table:
            return df_data

        data[sheet_name] = df_data

    wb.close()

    return data


class ReadXlsx(object):

    def __init__(self, *args, **kwargs):
        """
        读取 Excel 文件的类 用 openpyxl 重写

        # 例子
        xlsx_file = r"E:\my - 数据\其他\20180731 阿玲 参考\申请单G1.5(1)(1).xlsx"

        # 读取大文件的时候添加 read_only=True, data_only=True 的时候不读取公式
        xlsx = ReadXlsx(xlsx_file, read_only=True, data_only=True)

        xlsx.debug = True
        # print(xlsx.sheet_names)

        # 处理所有列表内容
        print('---------------------------------')
        while 1:
            values = xlsx.values()
            titles = next(values)
            print(xlsx.title, titles)

            # values 返回的是迭代器
            for line in values:
                data = {titles[i]: v for i, v in enumerate(line)}
                print(xlsx.line, line, data)
            if xlsx.next_sheet() is None:
                break
        print('---------------------------------')

        print(xlsx.set_sheet(10))  # 这个要判断下是否成功
        print(xlsx.next_sheet(), xlsx.index)
        """
        # print('args', args,'kwargs',kwargs)

        # 打开的表索引
        try:
            self.index = kwargs.pop('index')
        except:
            self.index = 0

        self.line = 0
        self.sheet = None

        # 转换表格格式
        self.conversion_format = True
        # 如果没有内容(None)的时候转换为空字符（需要设置 self.conversion_format = True）
        self.none_to_null_characters = True

        # 数字转字符串

        # 日期格式
        self.time_fmt = "%Y-%m-%d"  # YYYY-MM-DD
        # self.time_fmt = "%Y-%m"

        # self.excel = 'xlsx'
        self.wb = load_workbook(*args, **kwargs)
        self.sheet_names = self.wb.sheetnames
        self.set_sheet(self.index)
        # print(self.excel)

        self.debug = False

    @property
    def title(self):
        """  标题 """
        return self.sheet_names[self.index]

    @property
    def max_row(self):
        """  行 """
        if self.debug: print(f"title {self.title} index {self.index} 行 {self.sheet.max_row}")
        return self.sheet.max_row

    @property
    def max_column(self):
        """  列 """
        if self.debug: print(f"title {self.title} index {self.index} 列 {self.sheet.max_column}")
        return self.sheet.max_column

    # @property
    def datasets(self):
        """
        sheet 的数据集（列表）
        xls 的表达式不能显示
        """
        return list(self.values())

    def values(self):
        """
        sheet 的数据集（列表）
        xlsx 空表格是 None
        xls  空表格是 ''
        xlsx 0 是 0
        xls  0 是 0.0
        xls 的表达式不能显示
        # 返回可迭代结果
        """
        self.line = 0
        # rows 是可迭代的生成器
        for row in self.sheet.rows:
            if self.conversion_format:
                row_vals = []
                # row_vals = [c.value for c in row]
                # print(row)
                for c in row:
                    # fill = c.fill
                    # fg_color = fill.start_color.rgb  # start_color fgColor
                    # if 'Values must be of type' in str(fg_color):
                    #     fg_color = '00000000'
                    # print(fg_color)
                    # bg_color = fill.end_color.rgb  # end_color bgColor
                    # print(bg_color)
                    # print(c.data_type, c.number_format, is_date_format(c.number_format), c.value)
                    if c.value is None and self.none_to_null_characters:
                        cell_value = ''
                    elif c.data_type == "n":
                        if c.number_format != "General" and is_date_format(c.number_format):
                            # print('number_format 日期',c.number_format,c.value)
                            cell_value = c.value.strftime(self.time_fmt)
                        else:
                            # cell_value = str(c.value)  # 数字转换为字符串
                            cell_value = c.value
                    elif c.data_type == "d":
                        cell_value = c.value.strftime(self.time_fmt)
                    else:
                        cell_value = c.value
                    row_vals.append(cell_value)
            else:
                row_vals = [c.value for c in row]
            # print(row_vals)
            yield row_vals
            self.line += 1

    def next_sheet(self):
        """ 下一个表 """
        return self.set_sheet(self.index + 1)

    def set_sheet(self, index=None):
        """ 设置表 """

        if index is not None:
            # 这里判断是否超过范围 调试了很久 ， 因为索引是从0开始的 啊 啊 啊 啊 啊
            if len(self.sheet_names) >= index + 1:
                self.index = index
            else:
                # WarningInfo = "访问的表 %s 不存在" % index
                # warn(WarningInfo)
                # if self.index == index: # 初始化设置的不存在的表
                #     self.index = 0
                return None

        # 通过名字获取 sheet
        self.sheet = self.wb[self.sheet_names[self.index]]

        return self.index


def save_excel_specified_rows(input_file, output_file, start_row=1, end_row=11):
    """
    从Excel文件中读取指定的行，并保存到新文件

    :param input_file: str，输入的Excel文件路径
    :param output_file: str，输出的Excel文件路径
    :param start_row: int，起始行（1为起始索引）
    :param end_row: int，结束行（包含在结果内的索引）
    """

    # 加载输入工作簿
    input_workbook = load_workbook(input_file)

    # 创建一个新的工作簿
    output_workbook = Workbook()
    # 删除默认创建的第一张空表
    del output_workbook['Sheet']

    for sheet_name in input_workbook.sheetnames:
        input_sheet = input_workbook[sheet_name]

        # 创建一个新的工作表到输出工作簿中
        output_sheet = output_workbook.create_sheet(title=sheet_name)

        # 复制指定行到新的工作表中
        for row_num in range(start_row, end_row + 1):
            for col_num, cell in enumerate(input_sheet[row_num], 1):
                output_sheet.cell(row=row_num - start_row + 1, column=col_num, value=cell.value)

    # 保存到新文件
    output_workbook.save(output_file)


def save_dir_excel_specified_rows(file_dir, dst_dir=None, num_rows=10):
    """
    读取文件夹中的Excel文件保存前10行到新表格
    """
    file_dir = Path(file_dir)

    if dst_dir is None:
        dst_dir = file_dir / f'前十行数据'
    elif isinstance(dst_dir, str):
        dst_dir = Path(dst_dir)

    if not dst_dir.exists():
        dst_dir.mkdir()

    for file in get_dir_files(file_dir, ext='.xlsx'):
        save_excel_specified_rows(input_file=file, output_file=dst_dir / f'{os.path.basename(file)} {num_rows}行.xlsx', start_row=1, end_row=num_rows + 1)


def doc():
    """
    打印模块说明文档
    """
    doc_text = """"""
    doc_text += '\n'
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=ReadXlsx)
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=get_title_style)
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=get_excel_info)

    print(doc_text)


if __name__ == "__main__":
    t1 = time()

    # 例子
    xlsx_file = r"E:\my - 数据\其他\20180731 阿玲 参考\申请单G1.5(1)(1).xlsx"
    # data_only=True 的时候不读取公式 , index=10
    # 读取大文件的时候添加 read_only=True
    # xlsx = ReadExcel(xlsx_file, data_only=True)

    xlsx = ReadXlsx(xlsx_file)
    # time_fmt = "%Y-%m"
    xlsx.debug = True
    # print(xlsx.sheet_names)

    for i in range(len(xlsx.sheet_names)):
        xlsx.max_row
        xlsx.max_column
        print('next_sheet ---------------------------------')
        xlsx.next_sheet()

    # print(xlsx.datasets())
    # for i in xlsx.datasets():
    #     print(i)

    # for i in xlsx.values():
    #     print(i)

    print(xlsx.set_sheet(10))  # 这个要判断下是否成功
    print(xlsx.next_sheet(), xlsx.index)

    print(xlsx.title)

    print('用时 %.2f 秒' % (time() - t1))
