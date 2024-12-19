import os
import hashlib
from collections import OrderedDict

from ilds.file import get_dir_files, get_file_md5

import pandas as pd
from openpyxl import load_workbook


def get_df_list(file, sheet_names=None, concat_columns=None, add_source_column=True, strict_mode=False, is_print=True):
    df_list = []

    with pd.ExcelFile(file) as excel:
        if sheet_names is None:
            sheet_names = excel.sheet_names
        else:
            # 确保Excel文件中提供的 sheet_names 存在
            invalid_sheets = set(sheet_names) - set(excel.sheet_names)
            if invalid_sheets:
                raise ValueError(f"Excel 文件中不存在以下表格：{invalid_sheets}")

        if is_print:
            print('读取的表薄名称列表：', sheet_names, 'Excel表薄列表', excel.sheet_names)
        # 通过表薄名字读取表单
        # data['总表'] = pd.read_excel(excel, sheet_name='总表')
        # 读取全部表
        for sheet_name in sheet_names:
            _df = pd.read_excel(excel, sheet_name=sheet_name, header=0)
            # _df = pd.read_excel(excel, sheet_name=0, header=0)
            if is_print:
                print('表薄名称：', sheet_name, '\t', '行数：', len(_df), '\t', '文件：', file)
            # if 'Metadatas' != sheet_name:
            #     print('跳过内容', sheet_name)
            #     continue
            if add_source_column:
                _df['本行来自'] = f"{os.path.split(file)[1]} - {sheet_name}"
            # 是否指定要合并的列
            # 严格模式下检查工作表列是否与concat_columns完全匹配
            if concat_columns is not None:
                # 检查指定列是否都在DataFrame中
                missing_columns = set(concat_columns) - set(_df.columns)
                if missing_columns:
                    raise ValueError(f"工作表 '{sheet_name}' 中缺少以下指定列：{missing_columns}")

                # 严格模式：检查是否有额外的列
                if strict_mode:
                    extra_columns = set(_df.columns) - set(concat_columns)
                    if extra_columns:
                        raise ValueError(f"工作表 '{sheet_name}' 包含未在指定标题中声明的额外列：{extra_columns}")

                _df = _df[concat_columns]
            df_list.append(_df)
    return df_list


def get_excel_data(file, sheet_names=None, columns=None, add_source_column=True, only_read_first_table=False, read_sheet_state=False,
                   is_print=True):
    """
    读取 Excel 数据

    返回数据：'file_name', 'index', 'sheet_name', 'sheet_names', 'count', 'columns', 'df'

    :param file: Excel 文件
    :param sheet_names: 指定要读取的表，参数是要读取的表名的列表
    :param columns: 指定要读取的列，我们会只读取这些数据，方便用来合并
    :param add_source_column: 添加内容来源
    :param only_read_first_table: 只读取第一个表格
    :param read_sheet_state: 读取表格的状态，这样就可以处理隐藏表格
    :param is_print: 打印读取信息
    :return: {'file_name', 'index', 'sheet_name', 'sheet_names', 'count', 'columns', 'df'}
    """
    data = OrderedDict()

    if read_sheet_state:
        try:
            wb = load_workbook(file, read_only=False)
            sheet_state_data = {name: wb[name].sheet_state for name in wb.sheetnames}
            # print(sheet_state_data)
            wb.close()
        except Exception as e:
            print('get_excel_data 读取表格状态失败', e)
            sheet_state_data = {}
    else:
        sheet_state_data = {}

    with pd.ExcelFile(file) as excel:
        if sheet_names is None:
            sheet_names = excel.sheet_names
        else:
            # 确保Excel文件中提供的 sheet_names 存在
            invalid_sheets = set(sheet_names) - set(excel.sheet_names)
            if invalid_sheets:
                raise ValueError(f"Excel 文件中不存在以下表格：{invalid_sheets}")

        if is_print:
            print('读取的表薄名称列表：', sheet_names, 'Excel表薄列表', excel.sheet_names)

        # 通过表薄名字读取表单
        # data['总表'] = pd.read_excel(excel, sheet_name='总表')
        # 读取全部表
        for index, sheet_name in enumerate(sheet_names):
            sheet_state = sheet_state_data.get(sheet_name, None)
            _df = pd.read_excel(excel, sheet_name=sheet_name, header=0)
            file_name = os.path.basename(file)
            count = len(_df)

            if is_print:
                print('表薄名称：', sheet_name, '\t', '行数：', count, '\t', '文件：', file)

            if add_source_column:
                _df['本行来自'] = f"{file_name} - {sheet_name}"

            # 是否指定要合并的列
            if columns is not None:
                _df = _df[columns]

            df_data = {
                'file_name': file_name,
                'index': index,
                'sheet_name': sheet_name,
                'sheet_names': sheet_names,
                'sheet_state': sheet_state,
                'count': count,
                'columns': list(_df.columns),
                'df': _df,
            }

            if only_read_first_table:
                data = df_data
                break

            data[sheet_name] = df_data

    return data


def generate_cache_filename(file_path, file_hash, **read_excel_kwargs):
    """
    根据文件路径、文件内容哈希和读取参数生成缓存文件名
    """
    base_name = os.path.splitext(file_path)[0]
    # 创建哈希对象
    hash_obj = hashlib.md5()
    # 更新哈希对象
    hash_obj.update(str(read_excel_kwargs).encode('utf-8'))
    hash_obj.update(file_hash.encode('utf-8'))
    # 完整的哈希值用于命名缓存文件
    hash_digest = hash_obj.hexdigest()
    cache_file = f"{base_name}_{hash_digest}.cache"
    return cache_file


def load_excel_with_cache(file_path, **read_excel_kwargs):
    """
    从缓存加载 Excel文件

    先根据文件内容和参数的哈希值生成缓存文件名，如果缓存文件存在直接读取，否则读取Excel文件并保存缓存
    """
    # 计算 Excel 文件的哈希值
    file_hash = get_file_md5(file_path)
    # 根据文件哈希和参数生成缓存文件名
    cache_file = generate_cache_filename(file_path, file_hash, **read_excel_kwargs)

    # 检查缓存文件是否存在
    if os.path.exists(cache_file):
        print(f"从缓存载入数据: {cache_file}")
        return pd.read_parquet(cache_file)

    # 如果缓存不存在或文件内容改变，则从 Excel 文件加载
    print(f"从 Excel 载入数据: {file_path}")
    df = pd.read_excel(file_path, dtype_backend='pyarrow', **read_excel_kwargs)

    # 保存 DataFrame 到缓存（包括文件哈希以供验证）
    df.to_parquet(cache_file)
    print(f"保存到缓存: {cache_file}")

    return df


def read_song_ids(file_path, sheet_name=0, column='歌曲ID'):
    """
    读取Excel文件中的列的内容为列表

    参数：
    file_path: str Excel文件的路径
    sheet_name: str 或 int 工作表名称或编号，默认是第一个工作表

    返回：
    list 歌曲ID列的内容列表
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 确保列在数据框中
        if column not in df.columns:
            raise ValueError(f"Excel文件中不包含“{column}”列")

        # 获取列的内容
        song_ids = df[column].dropna().tolist()

        return song_ids
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return []


def read_excel(excel_file, sheet_name=None, dtype_backend='pyarrow'):
    """

    pandas sheet_name
    未指定（默认为 0）：将默认读取 Excel 文件中的第一个工作表（索引为 0 的工作表）
    支持指定表索引号、表名、表名列表
    None的时候，读取所有工作表，返回一个字典
    """
    # 读取所有工作表，返回一个字典
    dfs_all = pd.read_excel(excel_file, sheet_name=sheet_name, dtype_backend=dtype_backend)
    # for sheet, df in dfs_all.items():
    #     print(f"工作表名称: {sheet}")
    #     print(df.head())
    return dfs_all
