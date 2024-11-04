# -*- coding: utf-8 -*-
#
# ---------------------------------------
#   程序：pd.py
#   版本：0.9
#   作者：lds
#   日期：2024-09-12
#   语言：Python 3.X
#   说明：pandas 常用的函数集合，TODO 添加一些小抄在这里！
# ---------------------------------------
import os
from pathlib import Path
from collections import OrderedDict
import hashlib

from colorama import Fore, Back, Style

from ilds.file import get_dir_files, get_file_md5
from openpyxl import load_workbook

try:
    import pandas as pd
except ImportError as e:
    print(Fore.RED + "注：导入 pandas 失败，请安装它: pip install pandas", Style.RESET_ALL)
    pass


def get_columns_index(df, columns):
    """获取列名称在 DataFrame 中的索引"""
    if isinstance(columns, str):
        columns = [columns]
    columns_index = []
    for col in columns:
        columns_index.append(df.columns.get_loc(col))
    # print('columns_index', columns_index)
    return columns_index


def get_df_list(file, sheet_names=None, concat_columns=None, add_source_column=True, strict_mode=False, is_print=True):
    df_list = []

    with pd.ExcelFile(file) as excel:
        if sheet_names is None:
            sheet_names = excel.sheet_names
        else:
            # 确保Excel文件中提供的 sheet_names 存在
            invalid_sheets = set(sheet_names) - set(excel.sheet_names)
            if invalid_sheets:
                raise ValueError(f"Excel 文件中不存在以下表格：{invalid_sheets}")

        if is_print:
            print('读取的表薄名称列表：', sheet_names, 'Excel表薄列表', excel.sheet_names)
        # 通过表薄名字读取表单
        # data['总表'] = pd.read_excel(excel, sheet_name='总表')
        # 读取全部表
        for sheet_name in sheet_names:
            _df = pd.read_excel(excel, sheet_name=sheet_name, header=0)
            # _df = pd.read_excel(excel, sheet_name=0, header=0)
            if is_print:
                print('表薄名称：', sheet_name, '\t', '行数：', len(_df), '\t', '文件：', file)
            # if 'Metadatas' != sheet_name:
            #     print('跳过内容', sheet_name)
            #     continue
            if add_source_column:
                _df['本行来自'] = f"{os.path.split(file)[1]} - {sheet_name}"
            # 是否指定要合并的列
            # 严格模式下检查工作表列是否与concat_columns完全匹配
            if concat_columns is not None:
                # 检查指定列是否都在DataFrame中
                missing_columns = set(concat_columns) - set(_df.columns)
                if missing_columns:
                    raise ValueError(f"工作表 '{sheet_name}' 中缺少以下指定列：{missing_columns}")

                # 严格模式：检查是否有额外的列
                if strict_mode:
                    extra_columns = set(_df.columns) - set(concat_columns)
                    if extra_columns:
                        raise ValueError(f"工作表 '{sheet_name}' 包含未在指定标题中声明的额外列：{extra_columns}")

                _df = _df[concat_columns]
            df_list.append(_df)
    return df_list


def get_excel_data(file, sheet_names=None, columns=None, add_source_column=True, only_read_first_table=False, read_sheet_state=False,
                   is_print=True):
    """
    读取 Excel 数据

    返回数据：'file_name', 'index', 'sheet_name', 'sheet_names', 'count', 'columns', 'df'

    :param file: Excel 文件
    :param sheet_names: 指定要读取的表，参数是要读取的表名的列表
    :param columns: 指定要读取的列，我们会只读取这些数据，方便用来合并
    :param add_source_column: 添加内容来源
    :param only_read_first_table: 只读取第一个表格
    :param read_sheet_state: 读取表格的状态，这样就可以处理隐藏表格
    :param is_print: 打印读取信息
    :return: {'file_name', 'index', 'sheet_name', 'sheet_names', 'count', 'columns', 'df'}
    """
    data = OrderedDict()

    if read_sheet_state:
        try:
            wb = load_workbook(file, read_only=False)
            sheet_state_data = {name: wb[name].sheet_state for name in wb.sheetnames}
            # print(sheet_state_data)
            wb.close()
        except Exception as e:
            print('get_excel_data 读取表格状态失败', e)
            sheet_state_data = {}
    else:
        sheet_state_data = {}

    with pd.ExcelFile(file) as excel:
        if sheet_names is None:
            sheet_names = excel.sheet_names
        else:
            # 确保Excel文件中提供的 sheet_names 存在
            invalid_sheets = set(sheet_names) - set(excel.sheet_names)
            if invalid_sheets:
                raise ValueError(f"Excel 文件中不存在以下表格：{invalid_sheets}")

        if is_print:
            print('读取的表薄名称列表：', sheet_names, 'Excel表薄列表', excel.sheet_names)

        # 通过表薄名字读取表单
        # data['总表'] = pd.read_excel(excel, sheet_name='总表')
        # 读取全部表
        for index, sheet_name in enumerate(sheet_names):
            sheet_state = sheet_state_data.get(sheet_name, None)
            _df = pd.read_excel(excel, sheet_name=sheet_name, header=0)
            file_name = os.path.basename(file)
            count = len(_df)

            if is_print:
                print('表薄名称：', sheet_name, '\t', '行数：', count, '\t', '文件：', file)

            if add_source_column:
                _df['本行来自'] = f"{file_name} - {sheet_name}"

            # 是否指定要合并的列
            if columns is not None:
                _df = _df[columns]

            df_data = {
                'file_name': file_name,
                'index': index,
                'sheet_name': sheet_name,
                'sheet_names': sheet_names,
                'sheet_state': sheet_state,
                'count': count,
                'columns': list(_df.columns),
                'df': _df,
            }

            if only_read_first_table:
                data = df_data
                break

            data[sheet_name] = df_data

    return data


def generate_cache_filename(file_path, file_hash, **read_excel_kwargs):
    """
    根据文件路径、文件内容哈希和读取参数生成缓存文件名
    """
    base_name = os.path.splitext(file_path)[0]
    # 创建哈希对象
    hash_obj = hashlib.md5()
    # 更新哈希对象
    hash_obj.update(str(read_excel_kwargs).encode('utf-8'))
    hash_obj.update(file_hash.encode('utf-8'))
    # 完整的哈希值用于命名缓存文件
    hash_digest = hash_obj.hexdigest()
    cache_file = f"{base_name}_{hash_digest}.cache"
    return cache_file


def load_excel_with_cache(file_path, **read_excel_kwargs):
    """
    从缓存加载 Excel文件

    先根据文件内容和参数的哈希值生成缓存文件名，如果缓存文件存在直接读取，否则读取Excel文件并保存缓存
    """
    # 计算 Excel 文件的哈希值
    file_hash = get_file_md5(file_path)
    # 根据文件哈希和参数生成缓存文件名
    cache_file = generate_cache_filename(file_path, file_hash, **read_excel_kwargs)

    # 检查缓存文件是否存在
    if os.path.exists(cache_file):
        print(f"从缓存载入数据: {cache_file}")
        return pd.read_parquet(cache_file)

    # 如果缓存不存在或文件内容改变，则从 Excel 文件加载
    print(f"从 Excel 载入数据: {file_path}")
    df = pd.read_excel(file_path, dtype_backend='pyarrow', **read_excel_kwargs)

    # 保存 DataFrame 到缓存（包括文件哈希以供验证）
    df.to_parquet(cache_file)
    print(f"保存到缓存: {cache_file}")

    return df


def merging_excel_sheet(file, sheet_names=None, concat_columns=None, add_source_column=True, strict_mode=False, is_print=True, **concat_kwargs):
    """
    合并 Excel 表薄内容

    :param file: 要合并文件的路径
    :param sheet_names: 要合并的表
    :param concat_columns: 指定要合并的列名列表
    :param add_source_column: 是否添加原始来源
    :param strict_mode: 添加一个严格模式的参数，限制要合并文件的标题
    :param is_print: 打印信息
    :return:
    """
    df_list = get_df_list(file=file, sheet_names=sheet_names, concat_columns=concat_columns, add_source_column=add_source_column, strict_mode=strict_mode,
                          is_print=is_print)
    count = sum([len(df) for df in df_list])
    df = pd.concat(df_list, **concat_kwargs)  # result
    print('合并数据行数：', len(df), '导入数据行数统计：', count)
    return df


def merging_excel_file_data(file_dir_or_file_list, ext='', sheet_names=None, concat_columns=None, add_source_column=True, strict_mode=False, is_print=True,
                            **concat_kwargs):
    """
    合并多个 Excel 文件内容

    :param file_dir_or_file_list: 合并文件的路径或者文件列表
    :param ext: 要合并的文件后缀名，默认是文件夹中的全部文件
    :param sheet_names: 要合并的表
    :param concat_columns: 指定要合并的列名列表
    :param add_source_column: 是否添加原始来源
    :param strict_mode: 添加一个严格模式的参数，限制要合并文件的标题
    :param is_print: 打印信息
    :return:
    """
    all_len = 0
    frames = []

    if isinstance(file_dir_or_file_list, list):
        file_list = file_dir_or_file_list
    else:
        file_list = get_dir_files(file_dir_or_file_list, ext)

    for file in file_list:
        if is_print:
            print('处理文件:', file)
        df_list = get_df_list(file=file, sheet_names=sheet_names, concat_columns=concat_columns, add_source_column=add_source_column, strict_mode=strict_mode,
                              is_print=is_print)
        all_len += sum([len(df) for df in df_list])
        frames.extend(df_list)

    _df = pd.concat(frames, **concat_kwargs)  # result
    if is_print:
        print('合并数据行数：', len(_df), '原始数据行数：', all_len)
    return _df


def split_excel_sheet(file, dst_dir=None):
    """
    拆分 Excel 表薄内容

    :param file: 要拆分的文件
    :param dst_dir: 拆分文件的保存目录
    :return:
    """
    file = Path(file)

    if dst_dir is None:
        dst_dir = file.parent / f'{file.stem} - 拆分'
    elif isinstance(dst_dir, str):
        dst_dir = Path(dst_dir)

    if not dst_dir.exists():
        dst_dir.mkdir()

    xlsx = pd.ExcelFile(file)
    sheet_names = xlsx.sheet_names
    print('sheet_names', sheet_names)

    for sheet_name in sheet_names:
        df = xlsx.parse(sheet_name)
        to_file = dst_dir / f"{sheet_name}{file.suffix}"
        print(f"保存表薄: {sheet_name}, 文件: {to_file}")
        with pd.ExcelWriter(to_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"拆分完成，拆分文件保存在 '{dst_dir}' 文件夹中。")


def writer_excel(obj, path, index=False, use_zip64=False):
    """
    保存 Excel 文件

    :param obj:
    :param path:
    :param index: 是否添加索引
    :param use_zip64:
    :return:
    """
    if isinstance(obj, dict):
        with pd.ExcelWriter(path) as writer:
            for sheet_name, df_data in obj.items():
                df_data.to_excel(writer, sheet_name=sheet_name, index=index)
            if use_zip64:
                writer.book.use_zip64()
    else:
        with pd.ExcelWriter(path) as writer:
            obj.to_excel(writer, sheet_name='Sheet1', index=index)
            if use_zip64:
                writer.book.use_zip64()


def to_parquet(df, save_file):
    """
    保存为parquet数据函数，是为了记录使用方法
    """
    # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_parquet.html
    # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.to_parquet.html
    # 我们可以保存pyarrow数据，读取文件以后保存到当前文件夹的_d_文件夹中
    # df = pd.read_excel(file, dtype_backend='pyarrow')
    df.to_parquet(save_file)
    # restored_df = pd.read_parquet('data.parquet')


def add_sorted_sequence_number(df, sort_column, ascending=False, sort_name='排序序号'):
    """
    根据指定列的值生成排序序号，并将序号合并回原始DataFrame中

    @param df: 输入的DataFrame
    @param sort_column: 用于排序序号的列名
    @param ascending: 排序方式 True 表示升序排列，从小到大；False表示降序排列，从大到小（默认）
    @param sort_name: 排序列的名字
    @return: 添加了排序序号的 pd.DataFrame
    """
    # 将指定列转换为数值类型，无法转换的值变为 NaN
    df[sort_column] = pd.to_numeric(df[sort_column], errors='coerce')

    # 创建一个排序后的DataFrame
    df_sorted = df[[sort_column]].sort_values(by=sort_column, ascending=ascending).reset_index()

    # 添加序号列
    df_sorted[sort_name] = range(1, len(df_sorted) + 1)

    # 合并'排序序号'回原始DataFrame
    df = df.merge(df_sorted[['index', sort_name]], left_index=True, right_on='index').drop(columns=['index'])

    return df


def doc():
    """
    打印模块说明文档
    """
    doc_text = """"""
    doc_text += '\n'
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=get_columns_index)
    # doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=)
    print(doc_text)


if __name__ == '__main__':
    # 记录运行时间 --------------------------------------------------
    from time import time, sleep

    start_time = t1 = time()

    doc()

    print('运行时间 %.2f 秒' % (time() - start_time))
