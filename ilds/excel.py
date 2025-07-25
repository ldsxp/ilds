﻿# -*- coding: utf-8 -*-
#
# ---------------------------------------
#   程序：excel.py
#   版本：1.0
#   作者：lds
#   日期：2024-10-08
#   语言：Python 3.X
#   说明：读取excel文件内容
# ---------------------------------------

import os
from time import time
from warnings import warn
from pathlib import Path
import re
import string
import zipfile
import xml.etree.ElementTree as ET

import requests

from ilds.file import get_dir_files

from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, is_date_format
from openpyxl.utils import column_index_from_string
from openpyxl.packaging.relationship import get_rels_path, get_dependents, get_rel
from openpyxl.styles import Font, Alignment, PatternFill

(
    XL_CELL_EMPTY,
    XL_CELL_TEXT,
    XL_CELL_NUMBER,
    XL_CELL_DATE,
    XL_CELL_BOOLEAN,
    XL_CELL_ERROR,
    XL_CELL_BLANK,
) = range(7)


# warn('我们把 excel_xlsx 重命名为 excel，以后会删除 excel_xlsx', DeprecationWarning, stacklevel=2)


def get_title_style(workbook, style_colour='#daeef3', is_bold=False):
    """
    获取自定义颜色的标题风格
    :param workbook:
    :param style_colour: 默认颜色为什么浅色
    :param is_bold: 是否设置粗体字体
    :return:
    """
    head_style = workbook.add_format()
    # 设置粗体
    if is_bold:
        head_style.set_bold()
    # 设置对齐
    head_style.set_align('center')
    head_style.set_align('vcenter')
    # 设置颜色
    head_style.set_bg_color(style_colour)
    # 设置边框
    head_style.set_border()
    # 调整行的高度以适应文本
    head_style.set_text_wrap()
    return head_style


def get_excel_info(file, only_read_first_table=False):
    """
    读取 Excel 信息

    :param file: Excel 文件
    :param only_read_first_table: 只读取第一个表格
    :return: {sheet_name: {'file_name', 'index', 'sheet_name', 'sheet_names', 'max_row', 'max_column', 'columns'}, }
    """
    data = {}

    file_name = os.path.basename(file)

    wb = load_workbook(file, read_only=True)
    sheet_names = wb.sheetnames

    for index, sheet_name in enumerate(sheet_names):
        ws = wb[sheet_name]

        # 获取第一行数据
        columns = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        df_data = {
            'file_name': file_name,
            'index': index,
            'sheet_name': sheet_name,
            'sheet_names': sheet_names,
            'sheet_state': ws.sheet_state,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'columns': list(columns),

        }

        # 我们添加 count 为了和以前获取信息相同，以后会去掉，用 max_row 替代
        # df_data['count'] = df_data['max_row']

        if only_read_first_table:
            return df_data

        data[sheet_name] = df_data

    wb.close()

    return data


def parse_id(text):
    # 使用正则表达式来匹配 DISPIMG ID
    match = re.search(r'_xlfn.DISPIMG\("([^"]+)"', str(text))
    if match:
        return match.group(1)
    return None


class SheetImageLoader:
    """
    加载工作表中的所有图像

    from openpyxl import load_workbook
    # from sheet_image_loader import SheetImageLoader

    wb = load_workbook('path_to_file.xlsx')
    sheet = wb['Sheet1']

    # 将工作表放入加载器
    image_loader = SheetImageLoader(sheet)

    # 并从指定单元格获取 Pillow 图像
    cell = 'A3'
    image_raw = image_loader.get(cell)
    image = Image.open(io.BytesIO(image_raw))
    image.show()

    # 检查单元格中是否有图像
    if image_loader.image_in('A4'):
        print("有图片!")

    # https://openpyxl.readthedocs.io
    参考 https://github.com/ultr4nerd/openpyxl-image-loader
    """

    def __init__(self, excel_file, wb, sheet):
        """
        加载工作表图像
        """
        self.excel_file = excel_file
        self.z = zipfile.ZipFile(self.excel_file, 'r')
        self.wb = wb
        self.sheet = sheet
        self._images = {}
        self._external_images = {}
        self._wps_images = {}
        self.is_wps = None
        self.cell_image = {}
        self.is_cell_image = False
        self.file_list = self.z.namelist()

        self.embed_repeat_data = {}
        self.link_repeat_data = {}
        self._load_images_from_sheet()
        self.load_cell_image()
        self.read_drawing_images()

    def _load_images_from_sheet(self):
        # print(f'_load_images_from_sheet:{self.sheet._images}')
        for image in self.sheet._images:
            row = image.anchor._from.row + 1
            col = string.ascii_uppercase[image.anchor._from.col]
            self._images[f'{col}{row}'] = {'type': 'data', 'image': image._data}

    def load_cell_image(self):
        xml_path = f'xl/cellimages.xml'
        rels_xml_path = get_rels_path(xml_path)

        not_files = []

        # 先检查xml文件是否存在
        if xml_path not in self.file_list:
            not_files.append(xml_path)

        if rels_xml_path not in self.file_list:
            not_files.append(rels_xml_path)

        if not_files:
            print(f'file_list:{self.file_list}')
            print(f'load_cell_image 没有找到文件：{not_files}')
            return

        namespaces = {
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData'
        }
        tag_name = 'etc:cellImage'

        link_data = {}

        with self.z.open(xml_path) as f:
            # 解析XML内容
            tree = ET.parse(f)
            root = tree.getroot()

            for cell_image in root.findall(tag_name, namespaces):
                pic = cell_image.find('xdr:pic', namespaces)
                nvPicPr = pic.find('xdr:nvPicPr', namespaces)
                cNvPr = nvPicPr.find('xdr:cNvPr', namespaces)
                blipFill = pic.find('xdr:blipFill', namespaces)
                blip = blipFill.find('a:blip', namespaces)

                pic_id = cNvPr.get('id')
                pic_name = cNvPr.get('name')
                r_embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                r_link = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}link')

                image_info = {
                    'id': pic_id,
                    'name': pic_name,
                    'embed': r_embed,
                    'link': r_link,
                }
                # print(image_info)

                if r_embed:
                    ...
                    if r_embed in link_data:
                        if r_embed not in self.embed_repeat_data:
                            self.embed_repeat_data[r_embed] = set()
                        self.embed_repeat_data[r_embed].add(link_data[r_embed]['pic_name'])
                        self.embed_repeat_data[r_embed].add(pic_name)

                    link_data[r_embed] = {'r_embed': r_embed, 'pic_name': pic_name, }
                elif r_link:
                    if r_link in link_data:
                        if r_link not in self.link_repeat_data:
                            self.link_repeat_data[r_link] = set()
                        self.link_repeat_data[r_link].add(link_data[r_link]['pic_name'])
                        self.link_repeat_data[r_link].add(pic_name)

                    link_data[r_link] = {'r_link': r_link, 'pic_name': pic_name, }

                # self.cell_image[f'{col}{row}'] = {'type': 'data', 'image': image._data}

        if self.embed_repeat_data:
            embed_repeat_infos = []
            for k, v in self.embed_repeat_data.items():
                embed_repeat_infos.append(f"r_embed {k} 重复 :{v}")
            raise ValueError('\n'.join(embed_repeat_infos))

        if self.link_repeat_data:
            embed_repeat_infos = []
            for k, v in self.link_repeat_data.items():
                embed_repeat_infos.append(f"r_link {k} 重复 :{v}")
            raise ValueError('\n'.join(embed_repeat_infos))

        # XML命名空间
        namespace = {'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        # print(get_drawing_rels(z, ))
        with self.z.open(rels_xml_path) as drawing:
            # 解析XML文件
            tree = ET.parse(drawing)
            root = tree.getroot()

            # 遍历所有Relationship元素
            for i, relationship in enumerate(root.findall('rel:Relationship', namespace)):
                r_id = relationship.get('Id')
                type_ = relationship.get('Type')
                target = relationship.get('Target')
                target_mode = relationship.get('TargetMode')

                # target_mode True 图片链接 None 图片文件

                # print('load_cell_image', i, f'Id: {r_id}, Type: {type_}, Target: {target}, TargetMode: {target_mode}')

                if r_id not in link_data:
                    print(f'没有找到 {r_id} 的图片关联数据', link_data)
                    continue

                if target_mode:
                    self.cell_image[link_data[r_id]['pic_name']] = {'type': 'url', 'image': target}
                else:
                    self.cell_image[link_data[r_id]['pic_name']] = {'type': 'file', 'image': target}

    def get_drawing_link(self, z, xml_file):
        """
        获取图片的关联数据

        :param z:
        :param xml_file:
        :return:
        """

        link_data = {}
        wps_data = {}

        with z.open(xml_file) as f:
            # 解析XML内容
            tree = ET.parse(f)
            root = tree.getroot()

            # 定义命名空间
            namespaces = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }
            tag_name = 'xdr:twoCellAnchor'

            # 查找所有的twoCellAnchor元素
            for i, anchor in enumerate(root.findall(tag_name, namespaces)):
                # edit_as = anchor.get('editAs')

                col_row = None
                col = None
                row = None
                pic_name = None
                r_link = None
                to_col = None
                to_row = None

                from_elem = anchor.find('xdr:from', namespaces)
                if from_elem is not None:
                    col = from_elem.find('xdr:col', namespaces).text
                    col_off = from_elem.find('xdr:colOff', namespaces).text
                    row = from_elem.find('xdr:row', namespaces).text
                    row_off = from_elem.find('xdr:rowOff', namespaces).text
                    col_row = f'{string.ascii_uppercase[int(col)]}{int(row) + 1}'
                    # print(f"From - Column: {col}, Column Offset: {col_off}, Row: {row}, Row Offset: {row_off} col_row: {col_row}")

                to_elem = anchor.find('xdr:to', namespaces)
                if to_elem is not None:
                    to_col = to_elem.find('xdr:col', namespaces).text
                    to_col_off = to_elem.find('xdr:colOff', namespaces).text
                    to_row = to_elem.find('xdr:row', namespaces).text
                    to_row_off = to_elem.find('xdr:rowOff', namespaces).text
                    to_col_row = f'{string.ascii_uppercase[int(col)]}{row}'
                    # print(f"To - Column: {to_col}, Column Offset: {to_col_off}, Row: {to_row}, Row Offset: {to_row_off}")

                pic_elem = anchor.find('xdr:pic', namespaces)
                if pic_elem is not None:
                    cNvPr = pic_elem.find('.//xdr:cNvPr', namespaces)
                    if cNvPr is not None:
                        pic_id = cNvPr.get('id')
                        pic_name = cNvPr.get('name')
                        # print(f"Picture ID: {pic_id}, Name: {pic_name}")

                    blip_elem = pic_elem.find('.//a:blip', namespaces)
                    # print(blip_elem)
                    if blip_elem is not None:
                        r_link = blip_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}link')
                        r_embed = blip_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        # print(r_link, r_embed)
                        if r_link:
                            if r_link in link_data:
                                _data = {'col_row': col_row, 'col': col, 'row': row, 'to_col': to_col, 'to_row': to_row, 'r_link': r_link, 'pic_name': pic_name, }
                                info = f"已经存在 r_link {r_link} 单元格 {col_row}\n{link_data[r_link]}\n{_data}"
                                raise ValueError(info)
                            link_data[r_link] = {'col_row': col_row, 'col': col, 'row': row, 'to_col': to_col, 'to_row': to_row, 'r_link': r_link,
                                                 'pic_name': pic_name, }
                        elif r_embed:
                            if r_embed in link_data:
                                raise ValueError(f"已经存在 r_embed {r_embed} 单元格 {col_row}")
                            link_data[r_embed] = {'col_row': col_row, 'col': col, 'row': row, 'to_col': to_col, 'to_row': to_row, 'r_embed': r_embed,
                                                  'pic_name': pic_name, }

                    # 保存ID名字用于找到WPS的图片
                    if pic_name and r_link:
                        if pic_name in wps_data:
                            raise ValueError(f"已经存在 pic_name {pic_name} 单元格 {col_row}")
                        wps_data[pic_name] = r_link

                # print(i, col_row)

        return {'link_data': link_data, 'wps_data': wps_data, }

    def drawing_rels_external_images(self, z, xml_path, rels_xml_path):
        # print(f'drawing_rels_external_images(xml_path:{xml_path}) is_wps {self.is_wps} rels_xml_path:{rels_xml_path}')

        not_files = []

        # 先检查xml文件是否存在
        if xml_path not in self.file_list:
            not_files.append(xml_path)

        if rels_xml_path not in self.file_list:
            not_files.append(rels_xml_path)

        if not_files:
            print(f'file_list:{self.file_list}')
            print(f'没有找到xml文件：{not_files}')
            return

        r = self.get_drawing_link(z, xml_file=xml_path)

        link_data = r['link_data']
        # print('get_drawing_link', r)

        # XML命名空间
        namespace = {'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        # print(get_drawing_rels(z, ))
        with z.open(rels_xml_path) as drawing:
            # 解析XML文件
            tree = ET.parse(drawing)
            root = tree.getroot()

            # 遍历所有Relationship元素
            for i, relationship in enumerate(root.findall('rel:Relationship', namespace)):
                r_id = relationship.get('Id')
                type_ = relationship.get('Type')
                target = relationship.get('Target')
                target_mode = relationship.get('TargetMode')

                # target_mode True 图片链接 None 图片文件

                # print(i, f'Id: {r_id}, Type: {type_}, Target: {target}, TargetMode: {target_mode}')

                if r_id not in link_data:
                    print(f'没有找到 {r_id} 的图片关联数据', link_data)
                    continue

                # print(f'col_row:{col_row} r_id:{r_id} target:{target} {link_data[r_id]}')
                # if self.is_wps:
                #     pic_name = link_data[r_id]['pic_name']
                #     if pic_name in self._wps_images:
                #         raise ValueError(f"单元格 {col_row} 已经存在图像 {r_id} {self._images[col_row]} >  {target}")
                #     if target_mode:
                #         self._wps_images[pic_name] = {'type': 'url', 'image': target}
                #     else:
                #         self._wps_images[pic_name] = {'type': 'wps', 'image': target}
                # else:
                col_row = link_data[r_id]['col_row']
                if target_mode:
                    if col_row in self._images:
                        # image = self.get(col_row)
                        # image.show()
                        # import time
                        # time.sleep(10)
                        raise ValueError(f"单元格 {col_row} 已经存在图像 {r_id} {self._images[col_row]} >  {target}")
                    elif col_row in self._external_images:
                        raise ValueError(f"单元格 {col_row} 已经存在图像 {r_id} {self._external_images[col_row]} >  {target}")

                    self._external_images[col_row] = {'type': 'url', 'image': target}
                else:
                    if col_row not in self._images:
                        self._images[col_row] = {'type': 'file', 'image': target}

    def read_drawing_images(self):
        """
        读取 drawings 文件中的图片
        """
        index = self.wb.sheetnames.index(self.sheet.title)

        # 读取docProps/app.xml文件
        if self.is_wps is None:
            with self.z.open('docProps/app.xml') as app_xml:
                content = app_xml.read().decode('utf-8')
                # 检查内容中特定的字符串标识符
                if 'Microsoft Excel' in content:
                    self.is_wps = False
                elif 'WPS Office' in content or 'Kingsoft Office' in content:
                    self.is_wps = True
                else:
                    self.is_wps = None
                # print(self.is_wps, content)

        if self.is_wps:
            # 我们虽然可以判断是否是wps保存的，但是他的图片还是分两种，一种是wps原生的还有一种是wps转为浮动图片的
            # xml_path = f'xl/cellimages.xml'
            # rels_xml_path = f'xl/_rels/cellimages.xml.rels'
            xml_path = f'xl/drawings/drawing1.xml'
            # rels_xml_path = f'xl/drawings/_rels/drawing1.xml.rels'
            # if get_rels_path(f'xl/cellimages.xml') in self.file_list:
            #     xml_path = f'xl/cellimages.xml'
            #     self.is_cell_image = True
            # elif get_rels_path(f'xl/drawings/drawing1.xml') in self.file_list:
            #     xml_path = f'xl/drawings/drawing1.xml'
            # else:
            #     print('wps文件没事没有找到 rels 文件')
            #     return
            self.drawing_rels_external_images(self.z, xml_path=xml_path, rels_xml_path=get_rels_path(xml_path))
        else:
            xml_path = f'xl/drawings/drawing{index + 1}.xml'
            self.drawing_rels_external_images(self.z, xml_path=xml_path, rels_xml_path=get_rels_path(xml_path))
        # try:
        #     self.drawing_rels_external_images(self.z, xml_path=xml_path)
        # except Exception as e:
        #     print(f'读取网络图片发生错误:{e},\nself._external_images:{self._external_images}')
        #     self._external_images = {}
        #     pass

    def image_in(self, cell):
        """
        检查指定单元格中是否有图像
        """
        cell_id = parse_id(cell)
        return cell in self._images or cell in self._external_images or cell in self._wps_images or cell in self.cell_image or cell_id in self.cell_image

    def _get_image(self, data) -> bytes:
        """
        根据图像数据类型返回 PIL 图像
        """
        if data['type'] == 'url':
            url = data['image']
            response = requests.get(url)
            if response.status_code == 200:
                return response.content
            raise ValueError(f"获取网络图片失败 {data['image']}。状态代码: {response.status_code}")

        if data['type'] == 'wps' or data['type'] == 'file':
            # filename = data['image'].replace('../media', 'xl/media')
            filename = f"xl/media/{os.path.basename(data['image'])}"
            with zipfile.ZipFile(self.excel_file, 'r') as zip_file:
                return zip_file.read(filename)

        return data['image']()

    def get(self, cell, is_print=False) -> bytes:
        """
        从单元格中获取图像数据
        """
        if is_print:
            print(f'self._images : {self._images}')
            print(f'self._external_images : {self._external_images}')
            print(f'self._wps_images : {self._wps_images}')
            print(f'self.cell_image : {self.cell_image}')
        if cell in self._images:
            return self._get_image(self._images[cell])
        elif cell in self._external_images:
            return self._get_image(self._external_images[cell])
        elif cell in self._wps_images:
            return self._get_image(self._wps_images[cell])
        elif cell in self.cell_image:
            return self._get_image(self.cell_image[cell])

        cell_id = parse_id(cell)
        if cell_id in self._wps_images:
            return self._get_image(self._wps_images[cell_id])
        elif cell_id in self.cell_image:
            return self._get_image(self.cell_image[cell_id])

        raise ValueError(f"单元格 {cell} 不包含图像")

    def save_image(self, cell, file_name):
        """
        保存图片

        :param cell:
        :param file_name:
        :return:
        """
        with open(file_name, 'wb') as f:
            f.write(self.get(cell))

    def close(self):
        """关闭打开的 ZipFile 文件"""
        self.z.close()


class Excel:

    def __init__(self, *args, **kwargs):
        """
        使用 openpyxl 读取 Excel 文件

        # 例子
        excel_file = r""
        # data_only=True 的时候不读取公式
        # 读取大文件的时候添加 read_only=True
        excel = Excel(excel_file, data_only=True)
        # 读取大文件的时候添加 read_only=True, data_only=True 的时候不读取公式
        # excel = Excel(excel_file, read_only=True, data_only=True)

        excel.debug = True
        print(excel.sheet_names)

        # 处理所有列表内容
        print('---------------------------------')
        for sheet_name in excel.sheet_names:
            print('sheet_name', sheet_name)
            excel.set_sheet(sheet_name)
            values = excel.values()
            titles = next(values)
            print(excel.sheet_name, titles)

            # values 返回的是迭代器
            for line in values:
                if not any(line):
                    print(f'忽略内容为空的行: {line}')
                    continue
                data = {titles[i]: v for i, v in enumerate(line)}
                print(excel.line, line, data)

        print('---------------------------------')
        """
        # print('args', args, 'kwargs', kwargs)

        self.line = 0  # 当前行数
        self.row_values = []  # 转换以后的当前行数据

        # 转换表格格式
        self.conversion_format = True
        # 如果没有内容(None)的时候转换为空字符（需要设置 self.conversion_format = True）
        self.none_to_null_characters = True
        # 默认读取链接的实际地址
        self.is_hyperlink_target = True

        # 数字转字符串

        # 日期格式
        self.time_fmt = "%Y-%m-%d"  # 年-月-日格式

        # self.excel = 'xlsx'
        # 加载工作簿
        if args or 'filename' in kwargs:
            try:
                self.wb = load_workbook(*args, **kwargs)
            except FileNotFoundError:
                # 如果文件不存在，创建一个新的工作簿
                self.wb = Workbook()
                print(f'{args} {kwargs} 文件不存在，创建新表格')
        else:
            self.wb = Workbook()
            print('创建新表格')

        # print(self.excel)
        # 获取当前活动工作表
        self.sheet = self.wb.active

        # 插入新列数据
        self.new_column_start_idx = None
        self.header_idx = None
        self.existing_headers = None

        # 调试模式
        self.debug = False

    @property
    def title(self):
        """ 获取当前表的名字（已弃用，使用 sheet_name 替代）"""
        warn(
            "title 属性已弃用，请使用 sheet_name 属性代替",
            DeprecationWarning,
            stacklevel=2
        )
        return self.sheet_name

    @property
    def sheet_names(self):
        """ 工作表的名称列表 """
        return self.wb.sheetnames

    @property
    def sheet_name(self):
        """ 获取当前表的名字 """
        return self.sheet.title

    @property
    def max_row(self):
        """ 获取最大行数 """
        if self.debug:
            print(f"sheet_name {self.sheet_name} 行 {self.sheet.max_row}")
        return self.sheet.max_row

    @property
    def max_column(self):
        """ 获取最大列数 """
        if self.debug:
            print(f"sheet_name {self.sheet_name} 列 {self.sheet.max_column}")
        return self.sheet.max_column

    # @property
    def datasets(self):
        """
        sheet 的数据集（列表）
        xls 的表达式不能显示
        """
        return list(self.values())

    def values(self):
        """
        sheet 的数据集（列表）

        返回可迭代结果
        """
        self.line = 0
        for row in self.sheet.iter_rows():
            if self.conversion_format:
                # 使用列表解析优化行值的提取
                row_values = [self._convert_cell(c) for c in row]
            else:
                row_values = [c.value for c in row]

            self.row_values = row_values
            yield row_values
            self.line += 1

    def _convert_cell(self, cell):
        """根据单元格的数据类型和格式转换单元格的值"""
        # print(cell.data_type, cell.number_format, is_date_format(cell.number_format), cell.value,cell)
        if self.is_hyperlink_target and cell.hyperlink and cell.hyperlink.target:  # 我们没有使用 hasattr(cell, 'hyperlink') 是为了支持强制读取链接
            return cell.hyperlink.target
        if cell.value is None and self.none_to_null_characters:
            return ''
        if cell.data_type == "n":
            return cell.value.strftime(self.time_fmt) if is_date_format(cell.number_format) else cell.value  # cell.number_format != "General" and
        if cell.data_type == "d":
            return cell.value.strftime(self.time_fmt)

        return cell.value

    def next_sheet(self):
        """ 切换到下一个表 """
        info = """
        切换到下一个表以前我们用于循环获取表格的内容，现在使用下面的代码：
        for sheet_name in excel.sheet_names:
            print('sheet_name', sheet_name)
            excel.set_sheet(sheet_name)
        """
        raise ValueError(info)

    def set_sheet(self, sheet_name):
        """
        根据索引或者表名设置当前使用的表

        因为我们支持新建表，所以这里建议使用表名
        """
        if isinstance(sheet_name, int):
            if 0 <= sheet_name < len(self.sheet_names):
                sheet_name = self.sheet_names[sheet_name]
            else:
                raise ValueError(f"索引 {sheet_name} 的表不包含在 {self.sheet_names}")

        # 通过名字获取 sheet
        self.sheet = self.wb[sheet_name]

    def create_sheet(self, title=None, index=None):
        """ 创建表 """
        self.sheet = self.wb.create_sheet(title=title, index=index)
        return self.sheet

    def delete_sheet(self, sheet_name):
        """ 删除表 """
        del self.wb[sheet_name]
        # 获取当前活动工作表
        self.sheet = self.wb.active

    def append(self, iterable):
        """ 在当前工作表的底部添加一行 """
        self.sheet.append(iterable)
        self.line += 1

    def get_header_style(self):
        """
        获取标题的格式
        我们可以通过覆盖他来自定义格式
        """
        header_style = {
            # 'font': Font(name='Arial', size=12, bold=True),
            # 'alignment': Alignment(horizontal='center', vertical='center'),
            'fill': PatternFill(start_color='b7d8eb', end_color='b7d8eb', fill_type='solid')
        }
        return header_style

    def insert_column_headers(self, headers, header_idx=1):
        """ 插入标题行到当前表格 """
        # 保存原表的最大列数 + 1，这个属性将用于插入新列数据
        self.new_column_start_idx = self.sheet.max_column + 1
        self.header_idx = header_idx
        # 获取当前标题行的内容
        self.existing_headers = [cell.value for cell in self.sheet[header_idx]]
        print(f"现有标题行内容: {self.existing_headers} ，行长度 {len(self.existing_headers)} 最大列数 {self.sheet.max_column}")

        header_style = self.get_header_style()

        # 将新的列标题插入到标题行
        for offset, header in enumerate(headers):
            col_idx = self.new_column_start_idx + offset
            cell = self.sheet.cell(row=header_idx, column=col_idx, value=header)
            # 应用样式到单元格
            if 'font' in header_style:
                cell.font = header_style['font']
            if 'alignment' in header_style:
                cell.alignment = header_style['alignment']
            if 'fill' in header_style:
                cell.fill = header_style['fill']

    def insert_row_data(self, row_idx, row_data):
        """ 插入行内容到新标题的列中 """
        for col_offset, cell_value in enumerate(row_data):
            current_col_idx = self.new_column_start_idx + col_offset
            self.sheet.cell(row=row_idx, column=current_col_idx, value=cell_value)

    def insert_data_rows_example(self):
        """
        示例：插入列数据

        excel.insert_data_rows_example()
        excel.save(f'{excel_file} 处理结果.xlsx')
        exit()
        """

        new_headers = ['内容', '信息', ]

        self.insert_column_headers(headers=new_headers)

        # 处理每一行的数据 (从第二行开始，因为第一行是标题行)
        for row_idx, row_cells in enumerate(self.sheet.iter_rows(min_row=self.header_idx + 1), start=self.header_idx + 1):
            row_values = [self._convert_cell(c) for c in row_cells]
            print(row_values)

            # 处理数据
            processed_values = [f'内容-{row_values[2]}', f'信息{row_idx}', ]

            # 插入处理后的数据
            self.insert_row_data(row_idx, processed_values)

    def save(self, file_name):
        """ 保存工作簿到文件 """
        self.wb.save(file_name)


# 因为我们支持写入了，所以修改了名字，并且保留了原来的名字
ReadExcel = Excel


class ExcelWriter:
    """
    写入数据到 Excel 文件

    我们使用

    # 使用上下文的例子：
    with ExcelWriter("文件.xlsx") as excel_writer:
        excel_writer.append(tities)  # 写入表头
        for data_row in 数据列表:
            excel_writer.append(data_row)  # 写入数据行
    """

    def __init__(self, filename, sheet_name='Sheet1'):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

    def append(self, row):
        """
        添加一行数据到 Excel 中

        :param row: 要添加的数据行
        """
        self.ws.append(row)

    def auto_adjust_column_width(self):
        """
        自动调整列宽以适应内容
        """
        for column in self.ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def save(self):
        """
        保存 Excel 文件，带有检查文件存在的功能
        """
        if os.path.exists(self.filename):
            print(f"警告: 文件 '{self.filename}' 已存在，将被覆盖")
        self.auto_adjust_column_width()
        self.wb.save(self.filename)

    def __enter__(self):
        # 进入上下文管理器时返回自身
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        # 退出上下文管理器时自动保存
        self.save()


def save_excel_specified_rows(input_file, output_file, start_row=1, end_row=11):
    """
    从Excel文件中读取指定的行，并保存到新文件

    :param input_file: str，输入的Excel文件路径
    :param output_file: str，输出的Excel文件路径
    :param start_row: int，起始行（1为起始索引）
    :param end_row: int，结束行（包含在结果内的索引）
    """

    # 加载输入工作簿
    input_workbook = load_workbook(input_file)

    # 创建一个新的工作簿
    output_workbook = Workbook()
    # 删除默认创建的第一张空表
    del output_workbook['Sheet']

    for sheet_name in input_workbook.sheetnames:
        input_sheet = input_workbook[sheet_name]

        # 创建一个新的工作表到输出工作簿中
        output_sheet = output_workbook.create_sheet(title=sheet_name)

        # 复制指定行到新的工作表中
        for row_num in range(start_row, end_row + 1):
            for col_num, cell in enumerate(input_sheet[row_num], 1):
                output_sheet.cell(row=row_num - start_row + 1, column=col_num, value=cell.value)

    # 保存到新文件
    output_workbook.save(output_file)


def save_dir_excel_specified_rows(file_dir, dst_dir=None, num_rows=10):
    """
    读取文件夹中的Excel文件保存前10行到新表格
    """
    file_dir = Path(file_dir)

    if dst_dir is None:
        dst_dir = file_dir / f'前十行数据'
    elif isinstance(dst_dir, str):
        dst_dir = Path(dst_dir)

    if not dst_dir.exists():
        dst_dir.mkdir()

    for file in get_dir_files(file_dir, ext='.xlsx'):
        save_excel_specified_rows(input_file=file, output_file=dst_dir / f'{os.path.basename(file)} {num_rows}行.xlsx', start_row=1, end_row=num_rows + 1)


def split_excel(excel_file, save_dir=None, save_sheets=None, exclude=None):
    """
    拆分 Excel
    """
    if save_sheets is None:
        # 加载Excel文件
        original_wb = load_workbook(filename=excel_file)

        # 获取所有的工作表名字
        save_sheets = original_wb.sheetnames

    # 循环保存表
    for sheet_to_keep in save_sheets:

        if exclude and sheet_to_keep in exclude:
            print(f'忽略保存工作表: {sheet_to_keep}')

        # 重新加载工作簿
        wb = load_workbook(filename=excel_file)

        # 获取所有的工作表名称
        sheets = wb.sheetnames

        # 删除除当前工作表外的其他工作表
        for sheet_name in sheets:
            if sheet_name != sheet_to_keep:
                std = wb[sheet_name]
                wb.remove(std)

        if save_dir:
            save_file = os.path.join(save_dir, f"{sheet_to_keep}.xlsx")
        else:
            save_file = f"{sheet_to_keep}.xlsx"

        # 保存为新文件
        wb.save(save_file)
        print(f"工作表 '{sheet_to_keep}' 保存为 '{save_file}'")


def write_to_excel(file_name, sheet_name, data, start_cell='A1'):
    try:
        # 尝试加载存在的工作簿
        workbook = load_workbook(file_name)
        print(f"载入文件: {file_name}")
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿
        workbook = Workbook()
        print(f"新建文件: {file_name}")

    # 如果指定的工作表不存在，添加一个新的工作表
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
        print(f"创建表: {sheet_name}")
    else:
        sheet = workbook[sheet_name]
        print(f"写入表: {sheet_name}")

    if start_cell is None:
        # 直接写入表格
        for row_data in data:
            sheet.append(row_data)
    else:
        # 将起始单元格的列和行转换为数字索引
        start_col = column_index_from_string(start_cell[0])
        start_row = int(start_cell[1:])
        # 将数据写入工作表
        for row_index, row_data in enumerate(data):
            for col_index, value in enumerate(row_data):
                sheet.cell(row=start_row + row_index, column=start_col + col_index, value=value)

    # 保存工作簿到文件
    workbook.save(file_name)

    print(f"写入数据到文件 {file_name} 的 {sheet_name} ，开始位置 {start_cell}")


def doc():
    """
    打印模块说明文档
    """
    doc_text = """"""
    doc_text += '\n'
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=ReadExcel)
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=get_title_style)
    doc_text += '{fun.__name__}{fun.__doc__}\n'.format(fun=get_excel_info)

    print(doc_text)


if __name__ == "__main__":
    t1 = time()

    # 例子
    excel_file = Path('..') / '-' / '测试文件.xlsx'

    # excel_file = r""
    # data_only=True 的时候不读取公式
    # 读取大文件的时候添加 read_only=True
    excel = Excel(excel_file, data_only=True)
    # 读取大文件的时候添加 read_only=True, data_only=True 的时候不读取公式
    # excel = ReadExcel(excel_file, read_only=True, data_only=True)

    excel.debug = True
    # print(excel.sheet_names)

    # excel.insert_data_rows_example()
    # excel.save(f'{excel_file} 处理结果.xlsx')
    # exit()

    # 处理所有列表内容
    print('---------------------------------')
    for sheet_name in excel.sheet_names:
        print('sheet_name', sheet_name)
        excel.set_sheet(sheet_name)
        values = excel.values()
        titles = next(values)
        print(excel.sheet_name, titles)

        # values 返回的是迭代器
        for line in values:
            data = {titles[i]: v for i, v in enumerate(line)}
            print(excel.line, line, data)

    print('---------------------------------')

    print('用时 %.2f 秒' % (time() - t1))
